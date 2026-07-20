"""Geração de planilhas corporativas premium da consulta por setor."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from services.setor_service import COLUNAS_LISTVIEW, COLUNAS_SITUACAO


def _aplicar_preenchimento(celula: Any, hex_cor: str) -> None:
    from openpyxl.styles import PatternFill

    celula.fill = PatternFill("solid", fgColor=hex_cor)


def _escrever_cabecalho_premium(
    aba: Any,
    resumo: dict[str, str],
    total: int,
    *,
    titulo_secao: str,
    subtitulo: str = "Consulta por Setor",
) -> int:
    """Escreve marca, card de metadados e título da seção. Retorna linha da tabela."""
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    azul = "102A56"
    azul_claro = "EDF2FA"
    azul_faixa = "D9E2EF"
    muted = "607089"
    texto = "172B4D"
    fino = Side(style="thin", color=azul_faixa)

    # Linha 1–2: marca
    aba.merge_cells("A1:H1")
    titulo = aba["A1"]
    titulo.value = "RH BRIDA"
    titulo.font = Font(name="Calibri", bold=True, size=18, color=azul)
    titulo.alignment = Alignment(vertical="center")

    aba.merge_cells("A2:H2")
    sub = aba["A2"]
    sub.value = subtitulo
    sub.font = Font(name="Calibri", size=11, color=muted)
    sub.alignment = Alignment(vertical="center")

    aba.row_dimensions[1].height = 26
    aba.row_dimensions[2].height = 18

    # Linha 3: separador visual
    for col in range(1, 9):
        cel = aba.cell(row=3, column=col, value="")
        _aplicar_preenchimento(cel, azul)
    aba.row_dimensions[3].height = 4

    # Linha 5–8: card 4×2 (rótulo + valor)
    campos = [
        ("SETOR", resumo.get("Setor", "")),
        ("DIRETOR/SÓCIO", resumo.get("Diretor/Sócio", "")),
        ("GERENTE", resumo.get("Gerente", "")),
        ("GESTOR", resumo.get("Gestor", "")),
        ("DATA / HORA", resumo.get("Data/Hora", "")),
        ("USUÁRIO", resumo.get("Usuário", "")),
        ("TOTAL DE COLABORADORES", str(total)),
        ("", ""),
    ]

    # Cada campo ocupa 2 colunas: pares em grade 8 colunas
    layout = [
        (5, 1, campos[0]),
        (5, 3, campos[1]),
        (5, 5, campos[2]),
        (5, 7, campos[3]),
        (7, 1, campos[4]),
        (7, 3, campos[5]),
        (7, 5, campos[6]),
        (7, 7, campos[7]),
    ]

    for linha_rotulo, col_inicio, (rotulo, valor) in layout:
        linha_valor = linha_rotulo + 1
        col_fim = col_inicio + 1

        aba.merge_cells(
            start_row=linha_rotulo,
            start_column=col_inicio,
            end_row=linha_rotulo,
            end_column=col_fim,
        )
        aba.merge_cells(
            start_row=linha_valor,
            start_column=col_inicio,
            end_row=linha_valor,
            end_column=col_fim,
        )

        cel_rotulo = aba.cell(row=linha_rotulo, column=col_inicio, value=rotulo)
        cel_rotulo.font = Font(name="Calibri", size=9, color=muted, bold=True)
        cel_rotulo.alignment = Alignment(vertical="center", horizontal="left")
        _aplicar_preenchimento(cel_rotulo, azul_claro)

        cel_valor = aba.cell(
            row=linha_valor,
            column=col_inicio,
            value=valor or "—",
        )
        cel_valor.font = Font(name="Calibri", size=12, color=texto, bold=True)
        cel_valor.alignment = Alignment(vertical="center", horizontal="left")
        _aplicar_preenchimento(cel_valor, azul_claro)

        for r in (linha_rotulo, linha_valor):
            for c in (col_inicio, col_fim):
                cel = aba.cell(row=r, column=c)
                _aplicar_preenchimento(cel, azul_claro)
                cel.border = Border(
                    left=fino,
                    right=fino,
                    top=fino,
                    bottom=fino,
                )

        aba.row_dimensions[linha_rotulo].height = 16
        aba.row_dimensions[linha_valor].height = 22

    # Linha 9: espaço
    aba.row_dimensions[9].height = 10

    # Linha 10: título da seção de dados
    aba.merge_cells("A10:H10")
    secao = aba["A10"]
    secao.value = titulo_secao
    secao.font = Font(name="Calibri", bold=True, size=12, color=azul)
    secao.alignment = Alignment(vertical="center")
    aba.row_dimensions[10].height = 20

    # Larguras base das 8 primeiras colunas do card (tabela pode ter mais)
    for indice in range(1, 9):
        aba.column_dimensions[get_column_letter(indice)].width = 14

    return 11  # linha onde começa o cabeçalho da tabela


def _escrever_tabela(
    aba: Any,
    inicio: int,
    colunas: tuple[str, ...],
    registros: list[dict[str, str]],
) -> None:
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    azul = "102A56"
    zebra = "F5F8FC"
    borda = Side(style="thin", color="DCE4EF")

    for indice, coluna in enumerate(colunas, start=1):
        celula = aba.cell(row=inicio, column=indice, value=coluna)
        celula.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _aplicar_preenchimento(celula, azul)
        celula.border = Border(left=borda, right=borda, top=borda, bottom=borda)

    aba.row_dimensions[inicio].height = 22

    for linha_idx, registro in enumerate(registros, start=inicio + 1):
        for coluna_idx, coluna in enumerate(colunas, start=1):
            celula = aba.cell(
                row=linha_idx,
                column=coluna_idx,
                value=registro.get(coluna, "Não informado"),
            )
            celula.font = Font(name="Calibri", size=10, color="172B4D")
            celula.alignment = Alignment(vertical="center", wrap_text=True)
            celula.border = Border(left=borda, right=borda, top=borda, bottom=borda)
            if (linha_idx - inicio) % 2 == 0:
                _aplicar_preenchimento(celula, zebra)

    ultima_linha = inicio + max(len(registros), 1)
    ultima_coluna = get_column_letter(len(colunas))
    aba.auto_filter.ref = f"A{inicio}:{ultima_coluna}{ultima_linha}"
    aba.freeze_panes = f"A{inicio + 1}"

    for indice, coluna in enumerate(colunas, start=1):
        largura = max(len(coluna), 12)
        for registro in registros:
            largura = max(largura, len(str(registro.get(coluna, ""))))
        aba.column_dimensions[get_column_letter(indice)].width = min(
            largura + 2,
            42,
        )


def gerar_excel_consulta_setor(
    resumo: dict[str, str],
    registros: list[dict[str, str]],
    situacao: list[dict[str, str]],
    *,
    colunas_listview: tuple[str, ...] | None = None,
    colunas_situacao: tuple[str, ...] | None = None,
    subtitulo: str = "Consulta por Setor",
    titulo_listview: str = "Colaboradores",
    titulo_situacao: str = "Situação dos Colaboradores",
) -> bytes:
    """Gera Excel premium com abas Colaboradores e Situação."""
    from openpyxl import Workbook

    cols_lv = colunas_listview or COLUNAS_LISTVIEW
    cols_sit = colunas_situacao or COLUNAS_SITUACAO

    workbook = Workbook()
    aba_colaboradores = workbook.active
    aba_colaboradores.title = "Colaboradores"
    inicio = _escrever_cabecalho_premium(
        aba_colaboradores,
        resumo,
        len(registros),
        titulo_secao=titulo_listview,
        subtitulo=subtitulo,
    )
    _escrever_tabela(aba_colaboradores, inicio, cols_lv, registros)

    aba_situacao = workbook.create_sheet("Situação")
    inicio_situacao = _escrever_cabecalho_premium(
        aba_situacao,
        resumo,
        len(situacao),
        titulo_secao=titulo_situacao,
        subtitulo=subtitulo,
    )
    _escrever_tabela(
        aba_situacao,
        inicio_situacao,
        cols_sit,
        situacao,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
