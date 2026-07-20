"""Testes da consulta por setor com dados fictícios."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from services.setor_service import (
    COLUNAS_BLOQUEADAS,
    aplicar_filtros_consulta,
    buscar_por_termo,
    calcular_indicadores,
    nome_arquivo_seguro,
    paginar_registros,
    preparar_base_filtros,
    preparar_consulta_setor,
    preparar_listview,
    preparar_registros_setor,
    preparar_situacao,
)
from utils.exportacao_excel import gerar_excel_consulta_setor
from utils.exportacao_pdf import gerar_pdf_consulta_setor


def _base_ficticia() -> pd.DataFrame:
    linhas = []
    for indice in range(1, 26):
        linhas.append(
            {
                "Descrição": "LOGÍSTICA",
                "Empregado": str(1000 + indice),
                "Nome": f"COLABORADOR LOGISTICA {indice:02d}",
                "Função": "Operador",
                "CPF": f"{indice:011d}"[:11],
                "Nascimento": "01/01/1990",
                "Admissão": "01/01/2020",
                "Tempo": "6 anos e 6 meses",
                "NOME_GESTOR": "CARLOS SOUZA",
                "Gerente": "JULIANA SILVA",
                "Diretor/Sócio": "MARINO PEDRESCHI JUNIOR",
                "HORÁRIO DE TRABALHO": "08:00 às 17:00",
                "PcD": "Não",
                "TIPO_DEFICIENCIA": "",
                "Status": "Ativo" if indice < 23 else "Afastado",
                "DATA_AFASTAMENTO": "01/06/2026" if indice >= 23 else "",
                "MOTIVO_AFASTAMENTO": "Médico" if indice >= 23 else "",
                "TIPO AFASTAMENTO": "Temporário" if indice >= 23 else "",
                "TIPO DESLIGAMENTO": "",
                "FERIAS": "",
                "INICIO_FERIAS": "01/07/2026" if indice == 10 else "",
                "FIM_FERIAS": "20/07/2026" if indice == 10 else "",
                "DIAS_FERIAS": 20 if indice == 10 else 0,
                "RETORNO": "20/07/2026" if indice == 10 else (
                    "15/07/2026" if indice >= 23 else ""
                ),
                "Cel_Cv_corporativo": "11999990000",
                "Estab": "NAO_DEVE_APARECER",
                "Razão Social": "NAO_DEVE_APARECER",
                "CNPJ": "00000000000000",
                "CEI": "NAO",
                "Local": "NAO_DEVE_APARECER",
            }
        )
    linhas.append(
        {
            "Descrição": "FINANCEIRO",
            "Empregado": "2001",
            "Nome": "ANA FINANCEIRO",
            "Função": "Analista",
            "CPF": "12345678901",
            "Nascimento": "10/10/1988",
            "Admissão": "10/10/2019",
            "Tempo": "",
            "NOME_GESTOR": "MARCOS LIMA",
            "Gerente": "PATRICIA ALVES",
            "Diretor/Sócio": "MARINO PEDRESCHI JUNIOR",
            "HORÁRIO DE TRABALHO": "09:00 às 18:00",
            "PcD": "Sim",
            "TIPO_DEFICIENCIA": "Visual",
            "Status": "Ativo",
            "DATA_AFASTAMENTO": "",
            "MOTIVO_AFASTAMENTO": "",
            "TIPO AFASTAMENTO": "",
            "TIPO DESLIGAMENTO": "",
            "FERIAS": "Não",
            "DIAS_FERIAS": 0,
            "RETORNO": "",
            "Cel_Cv_corporativo": "11988887777",
            "Estab": "X",
            "Razão Social": "X",
            "CNPJ": "1",
            "CEI": "1",
            "Local": "X",
        }
    )
    return pd.DataFrame(linhas)


def test_preparar_base_filtros_cria_colunas_normalizadas() -> None:
    base = preparar_base_filtros(_base_ficticia())
    for coluna in (
        "setor",
        "diretor_socio",
        "gerente",
        "gestor",
        "grupo_cargo",
        "funcao",
    ):
        assert coluna in base.columns
    # Colunas originais preservadas.
    assert "Descrição" in base.columns
    assert "Gerente" in base.columns
    assert "NOME_GESTOR" in base.columns
    assert "Diretor/Sócio" in base.columns
    assert "Função" in base.columns
    # Sem AGRUP na base fictícia → "Não informado".
    assert set(base["grupo_cargo"]) == {"Não informado"}
    assert "LOGÍSTICA" in set(base["setor"])
    assert "JULIANA SILVA" in set(base["gerente"])
    assert "CARLOS SOUZA" in set(base["gestor"])
    assert "MARINO PEDRESCHI JUNIOR" in set(base["diretor_socio"])
    assert "Operador" in set(base["funcao"])


def test_aplicar_filtros_consulta_combina_com_busca() -> None:
    base = preparar_base_filtros(_base_ficticia())

    # Filtro por setor reduz a base antes da consulta textual.
    so_financeiro = aplicar_filtros_consulta(base, {"setor": ["FINANCEIRO"]})
    assert set(so_financeiro["setor"]) == {"FINANCEIRO"}
    assert len(so_financeiro) == 1

    # Filtro por gestor.
    so_gestor = aplicar_filtros_consulta(base, {"gestor": ["CARLOS SOUZA"]})
    assert set(so_gestor["gestor"]) == {"CARLOS SOUZA"}
    assert len(so_gestor) == 25

    # Filtro por gerente.
    so_gerente = aplicar_filtros_consulta(base, {"gerente": ["PATRICIA ALVES"]})
    assert set(so_gerente["gerente"]) == {"PATRICIA ALVES"}
    assert len(so_gerente) == 1

    # Sem seleção → base intacta.
    intacta = aplicar_filtros_consulta(base, {})
    assert len(intacta) == len(base)


def test_consulta_setor_apos_filtro_de_gestor() -> None:
    base = preparar_base_filtros(_base_ficticia())
    filtrada = aplicar_filtros_consulta(base, {"gestor": ["CARLOS SOUZA"]})
    consulta = preparar_consulta_setor(
        filtrada,
        "LOGÍSTICA",
        setor_selecionado="LOGÍSTICA",
    )
    assert consulta["estado"] == "resultados"
    assert consulta["paginacao"]["total_registros"] == 25
    assert consulta["resumo"]["Gestor"] == "CARLOS SOUZA"


def test_consulta_somente_com_filtro_sem_texto() -> None:
    """Filtro sozinho (sem termo) deve executar a consulta."""
    base = preparar_base_filtros(_base_ficticia())
    filtrada = aplicar_filtros_consulta(base, {"setor": ["FINANCEIRO"]})
    consulta = preparar_consulta_setor(filtrada, "")
    assert consulta["estado"] == "resultados"
    assert consulta["setor"] == "FINANCEIRO"
    assert consulta["paginacao"]["total_registros"] == 1


def test_consulta_filtro_sem_texto_multiplos_setores() -> None:
    """Filtro amplo sem termo → pede seleção quando há vários setores."""
    base = preparar_base_filtros(_base_ficticia())
    # Gerente da LOGÍSTICA só existe nela; gestor comum + termo vazio
    # com setor Todos ainda cobre 2 setores se não filtrarmos.
    # Usando grupo_cargo (todos NI) não restringe; filtramos por gestor
    # que só está na LOGÍSTICA + FINANCEIRO tem outro gestor.
    filtrada = aplicar_filtros_consulta(
        base,
        {"gestor": ["CARLOS SOUZA", "MARCOS LIMA"]},
    )
    consulta = preparar_consulta_setor(filtrada, "")
    assert consulta["estado"] == "selecionar_setor"
    assert len(consulta["setores"]) >= 2


def test_consulta_busca_e_filtro_combinados() -> None:
    base = preparar_base_filtros(_base_ficticia())
    filtrada = aplicar_filtros_consulta(base, {"gerente": ["JULIANA SILVA"]})
    consulta = preparar_consulta_setor(
        filtrada,
        "LOGÍSTICA",
        setor_selecionado="LOGÍSTICA",
    )
    assert consulta["estado"] == "resultados"
    assert consulta["paginacao"]["total_registros"] == 25


def test_filtros_consulta_setor_composicao() -> None:
    from components.dashboard_filters import campos_filtro_submenu

    assert campos_filtro_submenu("Consulta por Setor") == (
        "setor",
        "diretor_socio",
        "gerente",
        "gestor",
        "grupo_cargo",
        "funcao",
    )
    assert "status" not in campos_filtro_submenu("Consulta por Setor")
    assert "genero" not in campos_filtro_submenu("Consulta por Setor")


def test_busca_por_setor_gerente_gestor_nome_e_matricula() -> None:
    dados = _base_ficticia()
    assert not buscar_por_termo(dados, "logistica").empty
    assert not buscar_por_termo(dados, "JULIANA").empty
    assert not buscar_por_termo(dados, "carlos souza").empty
    assert not buscar_por_termo(dados, "ANA FIN").empty
    assert buscar_por_termo(dados, "2001").iloc[0]["Nome"] == "ANA FINANCEIRO"


def test_busca_vazia_nao_retorna_todos() -> None:
    assert buscar_por_termo(_base_ficticia(), "   ").empty


def test_consulta_sem_resultado() -> None:
    consulta = preparar_consulta_setor(_base_ficticia(), "inexistente-xyz")
    assert consulta["estado"] == "sem_resultados"


def test_selecao_quando_multiplos_setores() -> None:
    consulta = preparar_consulta_setor(_base_ficticia(), "a")
    assert consulta["estado"] == "selecionar_setor"
    assert len(consulta["setores"]) >= 2


def test_apos_setor_lista_todos_colaboradores_do_setor() -> None:
    consulta = preparar_consulta_setor(
        _base_ficticia(),
        "COLABORADOR LOGISTICA 01",
        setor_selecionado="LOGÍSTICA",
    )
    assert consulta["estado"] == "resultados"
    assert consulta["paginacao"]["total_registros"] == 25


def test_paginacao_vinte_por_pagina() -> None:
    registros = preparar_registros_setor(
        _base_ficticia().query("Descrição == 'LOGÍSTICA'"),
        referencia=date(2026, 7, 14),
    )
    pagina_1 = paginar_registros(registros, 1)
    pagina_2 = paginar_registros(registros, 2)
    assert len(pagina_1["registros"]) == 20
    assert len(pagina_2["registros"]) == 5
    assert pagina_1["total_paginas"] == 2
    assert pagina_1["tem_anterior"] is False
    assert pagina_1["tem_proxima"] is True


def test_ordenacao_alfabetica_e_listview_sem_colunas_bloqueadas() -> None:
    registros = preparar_registros_setor(_base_ficticia())
    nomes = [item["Nome"] for item in registros]
    assert nomes == sorted(nomes, key=str.casefold)
    tabela = preparar_listview(registros[:5])
    for coluna in COLUNAS_BLOQUEADAS:
        assert coluna not in tabela.columns
    assert list(tabela.columns)[:3] == ["Nome", "Idade", "Cargo"]
    assert "Idade" in registros[0]
    assert all(valor == "***.***.***-**" or valor == "Não informado" for valor in tabela["CPF"])


def test_indicadores_e_situacao() -> None:
    consulta = preparar_consulta_setor(
        _base_ficticia(),
        "LOGÍSTICA",
        consulta_em=datetime(2026, 7, 14, 14, 52),
        referencia=date(2026, 7, 14),
    )
    indicadores = calcular_indicadores(
        consulta["registros"], referencia=date(2026, 7, 14)
    )
    assert indicadores["Total de colaboradores"] == 25
    assert indicadores["Afastados"] == 3
    assert indicadores["Em férias"] == 1
    situacao = preparar_situacao(consulta["registros"])
    assert not situacao.empty
    assert "Data de Afastamento" in situacao.columns
    assert consulta["resumo"]["Usuário"] == "Não identificado"
    assert "Local" not in consulta["resumo"]


def test_kpi_ferias_nao_conta_agendamento_futuro() -> None:
    """Agendada (Marcada) no futuro não entra no cartão Em férias."""
    registros = [
        {
            "Status": "Ativo",
            "Férias": "Sem férias",
            "Início Férias": "Não informado",
            "Fim Férias": "Não informado",
        },
        {
            "Status": "Ativo",
            "Férias": "Marcada · 24/09/2026 às 12/10/2026",
            "Início Férias": "24/09/2026",
            "Fim Férias": "12/10/2026",
        },
    ]
    indicadores = calcular_indicadores(
        registros, referencia=date(2026, 7, 20)
    )
    assert indicadores["Em férias"] == 0
    assert indicadores["Ativos"] == 2


def test_nome_arquivo_seguro() -> None:
    nome = nome_arquivo_seguro(
        "LOGÍSTICA / AÇÚCAR",
        "pdf",
        datetime(2026, 7, 14, 10, 35, 0),
    )
    assert nome.startswith("consulta_setor_")
    assert nome.endswith(".pdf")
    assert "/" not in nome
    assert " " not in nome


def test_exportacoes_pdf_e_excel() -> None:
    consulta = preparar_consulta_setor(_base_ficticia(), "LOGÍSTICA")
    registros = consulta["registros"]
    situacao = preparar_situacao(registros).to_dict(orient="records")
    pdf = gerar_pdf_consulta_setor(consulta["resumo"], registros, situacao)
    excel = gerar_excel_consulta_setor(
        consulta["resumo"],
        registros,
        situacao,
    )
    assert pdf.startswith(b"%PDF")
    planilha = load_workbook(BytesIO(excel))
    assert planilha.sheetnames == ["Colaboradores", "Situação"]
    aba = planilha["Colaboradores"]
    assert aba["A1"].value == "RH BRIDA"
    assert aba["A2"].value == "Consulta por Setor"
    assert aba.cell(row=5, column=1).value == "SETOR"
    assert aba.cell(row=5, column=3).value == "DIRETOR/SÓCIO"
    assert aba.cell(row=6, column=1).value  # valor do setor
    assert aba.cell(row=10, column=1).value == "Colaboradores"
    # Cabeçalho da tabela começa na linha 11
    assert aba.cell(row=11, column=1).value == "Nome"
    cabecalhos = [
        aba.cell(row=11, column=c).value for c in range(1, 15)
    ]
    assert "Diretor/Sócio" in cabecalhos
    valores = " ".join(
        str(valor or "")
        for linha in aba.iter_rows(values_only=True)
        for valor in linha
    )
    for bloqueada in COLUNAS_BLOQUEADAS:
        assert bloqueada not in valores
    assert aba.auto_filter.ref
    assert aba.freeze_panes


def _base_hierarquia_realistica() -> pd.DataFrame:
    """Gerente reporta a sócio; gestores reportam ao gerente; equipes nos gestores."""
    return pd.DataFrame(
        [
            {
                "Descrição": "LOGÍSTICA",
                "Empregado": "1",
                "Nome": "FABIO HENRIQUE SGOBI",
                "Função": "SOCIO",
                "NOME_GESTOR": "FABIO HENRIQUE SGOBI",
                "Gerente": "FABIO HENRIQUE SGOBI",
                "AGRUP_CARGOS_FUNCOES": "DIRETOR",
                "Status": "Ativo",
                "FERIAS": "Não",
                "PcD": "Não",
                "TIPO_DEFICIENCIA": "",
                "HORÁRIO DE TRABALHO": "08:00",
                "CPF": "00000000001",
                "Admissão": "01/01/2000",
                "Tempo": "20",
                "Cel_Cv_corporativo": "",
                "Estab": "X",
                "Razão Social": "X",
                "CNPJ": "1",
                "CEI": "1",
                "Local": "1",
            },
            {
                "Descrição": "LOGÍSTICA",
                "Empregado": "2",
                "Nome": "MARCELO JOSE PRIOR",
                "Função": "GERENTE DE LOGISTICA",
                "NOME_GESTOR": "FABIO HENRIQUE SGOBI",
                "Gerente": "FABIO HENRIQUE SGOBI",
                "AGRUP_CARGOS_FUNCOES": "LOGISTICA",
                "Status": "Ativo",
                "FERIAS": "Não",
                "PcD": "Não",
                "TIPO_DEFICIENCIA": "",
                "HORÁRIO DE TRABALHO": "08:00",
                "CPF": "00000000002",
                "Admissão": "01/01/2010",
                "Tempo": "10",
                "Cel_Cv_corporativo": "",
                "Estab": "X",
                "Razão Social": "X",
                "CNPJ": "1",
                "CEI": "1",
                "Local": "1",
            },
            {
                "Descrição": "LOGÍSTICA",
                "Empregado": "3",
                "Nome": "ARTUR CARNEIRO FERREIRA",
                "Função": "SUPERVISOR DE LOGISTICA",
                "NOME_GESTOR": "MARCELO JOSE PRIOR",
                "Gerente": "MARCELO JOSE PRIOR",
                "AGRUP_CARGOS_FUNCOES": "LOGISTICA",
                "Status": "Ativo",
                "FERIAS": "Não",
                "PcD": "Não",
                "TIPO_DEFICIENCIA": "",
                "HORÁRIO DE TRABALHO": "08:00",
                "CPF": "00000000003",
                "Admissão": "01/01/2015",
                "Tempo": "5",
                "Cel_Cv_corporativo": "",
                "Estab": "X",
                "Razão Social": "X",
                "CNPJ": "1",
                "CEI": "1",
                "Local": "1",
            },
            {
                "Descrição": "LOGÍSTICA",
                "Empregado": "4",
                "Nome": "DENER BENICIO DELGADO",
                "Função": "SUPERVISOR DE CARGA E DESCARGA",
                "NOME_GESTOR": "ARTUR CARNEIRO FERREIRA",
                "Gerente": "MARCELO JOSE PRIOR",
                "AGRUP_CARGOS_FUNCOES": "LOGISTICA",
                "Status": "Ativo",
                "FERIAS": "Não",
                "PcD": "Não",
                "TIPO_DEFICIENCIA": "",
                "HORÁRIO DE TRABALHO": "08:00",
                "CPF": "00000000004",
                "Admissão": "01/01/2016",
                "Tempo": "4",
                "Cel_Cv_corporativo": "",
                "Estab": "X",
                "Razão Social": "X",
                "CNPJ": "1",
                "CEI": "1",
                "Local": "1",
            },
            {
                "Descrição": "LOGÍSTICA",
                "Empregado": "5",
                "Nome": "COLAB ARTUR 1",
                "Função": "MOTORISTA",
                "NOME_GESTOR": "ARTUR CARNEIRO FERREIRA",
                "Gerente": "MARCELO JOSE PRIOR",
                "AGRUP_CARGOS_FUNCOES": "LOGISTICA",
                "Status": "Ativo",
                "FERIAS": "Não",
                "PcD": "Não",
                "TIPO_DEFICIENCIA": "",
                "HORÁRIO DE TRABALHO": "08:00",
                "CPF": "00000000005",
                "Admissão": "01/01/2020",
                "Tempo": "2",
                "Cel_Cv_corporativo": "",
                "Estab": "X",
                "Razão Social": "X",
                "CNPJ": "1",
                "CEI": "1",
                "Local": "1",
            },
            {
                "Descrição": "LOGÍSTICA",
                "Empregado": "6",
                "Nome": "COLAB DENER 1",
                "Função": "AJUDANTE",
                "NOME_GESTOR": "DENER BENICIO DELGADO",
                "Gerente": "MARCELO JOSE PRIOR",
                "AGRUP_CARGOS_FUNCOES": "LOGISTICA",
                "Status": "Ativo",
                "FERIAS": "Não",
                "PcD": "Não",
                "TIPO_DEFICIENCIA": "",
                "HORÁRIO DE TRABALHO": "08:00",
                "CPF": "00000000006",
                "Admissão": "01/01/2021",
                "Tempo": "1",
                "Cel_Cv_corporativo": "",
                "Estab": "X",
                "Razão Social": "X",
                "CNPJ": "1",
                "CEI": "1",
                "Local": "1",
            },
            {
                "Descrição": "LOGÍSTICA",
                "Empregado": "7",
                "Nome": "COLAB DENER 2",
                "Função": "AJUDANTE",
                "NOME_GESTOR": "DENER BENICIO DELGADO",
                "Gerente": "MARCELO JOSE PRIOR",
                "AGRUP_CARGOS_FUNCOES": "LOGISTICA",
                "Status": "Ativo",
                "FERIAS": "Não",
                "PcD": "Não",
                "TIPO_DEFICIENCIA": "",
                "HORÁRIO DE TRABALHO": "08:00",
                "CPF": "00000000007",
                "Admissão": "01/01/2022",
                "Tempo": "1",
                "Cel_Cv_corporativo": "",
                "Estab": "X",
                "Razão Social": "X",
                "CNPJ": "1",
                "CEI": "1",
                "Local": "1",
            },
            {
                "Descrição": "ADM",
                "Empregado": "8",
                "Nome": "OUTRO SETOR",
                "Função": "ANALISTA",
                "NOME_GESTOR": "OUTRO GESTOR",
                "Gerente": "OUTRO GERENTE",
                "AGRUP_CARGOS_FUNCOES": "ADM",
                "Status": "Ativo",
                "FERIAS": "Não",
                "PcD": "Não",
                "TIPO_DEFICIENCIA": "",
                "HORÁRIO DE TRABALHO": "08:00",
                "CPF": "00000000008",
                "Admissão": "01/01/2020",
                "Tempo": "2",
                "Cel_Cv_corporativo": "",
                "Estab": "X",
                "Razão Social": "X",
                "CNPJ": "1",
                "CEI": "1",
                "Local": "1",
            },
        ]
    )


def test_filtro_gestor_inclui_proprio_gestor() -> None:
    base = preparar_base_filtros(_base_hierarquia_realistica())
    filtrada = aplicar_filtros_consulta(base, {"gestor": ["DENER BENICIO DELGADO"]})
    nomes = set(filtrada["Nome"].tolist())
    assert "DENER BENICIO DELGADO" in nomes
    assert "COLAB DENER 1" in nomes
    assert "COLAB DENER 2" in nomes
    assert "ARTUR CARNEIRO FERREIRA" not in nomes
    assert len(filtrada) == 3
    assert len(filtrada) == len(filtrada.drop_duplicates(subset=["Empregado"]))


def test_filtro_gerente_inclui_gestores_equipe_e_proprio() -> None:
    base = preparar_base_filtros(_base_hierarquia_realistica())
    filtrada = aplicar_filtros_consulta(base, {"gerente": ["MARCELO JOSE PRIOR"]})
    nomes = set(filtrada["Nome"].tolist())

    assert "MARCELO JOSE PRIOR" in nomes
    assert "DENER BENICIO DELGADO" in nomes
    assert "ARTUR CARNEIRO FERREIRA" in nomes
    assert "COLAB ARTUR 1" in nomes
    assert "COLAB DENER 1" in nomes
    assert "COLAB DENER 2" in nomes
    assert "FABIO HENRIQUE SGOBI" not in nomes
    assert "OUTRO SETOR" not in nomes
    assert len(filtrada) == 6
    assert len(filtrada) == len(filtrada.drop_duplicates(subset=["Empregado"]))


def test_filtro_gerente_mesmo_resultado_para_kpi_e_registros() -> None:
    base = preparar_base_filtros(_base_hierarquia_realistica())
    filtrada = aplicar_filtros_consulta(base, {"gerente": ["MARCELO JOSE PRIOR"]})
    consulta = preparar_consulta_setor(filtrada, "")
    assert consulta["estado"] == "resultados"
    assert consulta["paginacao"]["total_registros"] == 6
    assert consulta["indicadores"]["Total de colaboradores"] == 6
    nomes = {r["Nome"] for r in consulta["registros"]}
    assert "DENER BENICIO DELGADO" in nomes
    assert "MARCELO JOSE PRIOR" in nomes


def test_filtro_gerente_unifica_multiplos_setores() -> None:
    """Guarda-chuva do gerente não pode sumir atrás de 'selecionar setor'."""
    base = preparar_base_filtros(_base_hierarquia_realistica())
    # Força dois setores sob o mesmo gerente.
    base.loc[base["Nome"] == "COLAB DENER 1", "Descrição"] = "MOTORISTA"
    base.loc[base["Nome"] == "COLAB DENER 1", "setor"] = "MOTORISTA"
    base.loc[base["Nome"] == "COLAB DENER 2", "Descrição"] = "MOTORISTA"
    base.loc[base["Nome"] == "COLAB DENER 2", "setor"] = "MOTORISTA"

    filtrada = aplicar_filtros_consulta(base, {"gerente": ["MARCELO JOSE PRIOR"]})
    sem_unificar = preparar_consulta_setor(filtrada, "")
    assert sem_unificar["estado"] == "selecionar_setor"

    consulta = preparar_consulta_setor(filtrada, "", unificar_setores=True)
    assert consulta["estado"] == "resultados"
    assert consulta["paginacao"]["total_registros"] == 6
    assert "DENER BENICIO DELGADO" in {r["Nome"] for r in consulta["registros"]}
    assert "MOTORISTA" in str(consulta["setor"])
    assert "LOGÍSTICA" in str(consulta["setor"]) or "LOGISTICA" in str(consulta["setor"])


def test_setor_selecionado_invalido_e_rejeitado() -> None:
    base = preparar_base_filtros(_base_ficticia())
    filtrada = aplicar_filtros_consulta(base, {"setor": ["FINANCEIRO"]})
    consulta = preparar_consulta_setor(
        filtrada,
        "",
        setor_selecionado="SETOR_INEXISTENTE",
    )
    assert consulta["estado"] == "resultados"
    assert consulta["setor"] == "FINANCEIRO"


def test_filtro_diretor_socio_guarda_chuva_e_diretos() -> None:
    """Diretor inclui estrutura + ligado direto sem gerente/gestor."""
    linhas = [
        {
            "Descrição": "DIR",
            "Empregado": "1",
            "Nome": "MARINO PEDRESCHI JUNIOR",
            "Função": "DIRETOR",
            "NOME_GESTOR": "",
            "Gerente": "",
            "Diretor/Sócio": "",
            "Status": "Ativo",
            "FERIAS": "Não",
        },
        {
            "Descrição": "ADM",
            "Empregado": "2",
            "Nome": "GERENTE SOB DIRETOR",
            "Função": "GERENTE",
            "NOME_GESTOR": "",
            "Gerente": "",
            "Diretor/Sócio": "MARINO PEDRESCHI JUNIOR",
            "Status": "Ativo",
            "FERIAS": "Não",
        },
        {
            "Descrição": "ADM",
            "Empregado": "3",
            "Nome": "GESTOR SOB GERENTE",
            "Função": "GESTOR",
            "NOME_GESTOR": "",
            "Gerente": "GERENTE SOB DIRETOR",
            "Diretor/Sócio": "MARINO PEDRESCHI JUNIOR",
            "Status": "Ativo",
            "FERIAS": "Não",
        },
        {
            "Descrição": "ADM",
            "Empregado": "4",
            "Nome": "COLAB SOB GESTOR",
            "Função": "ANALISTA",
            "NOME_GESTOR": "GESTOR SOB GERENTE",
            "Gerente": "GERENTE SOB DIRETOR",
            "Diretor/Sócio": "MARINO PEDRESCHI JUNIOR",
            "Status": "Ativo",
            "FERIAS": "Não",
        },
        {
            "Descrição": "ADM",
            "Empregado": "5",
            "Nome": "COLAB DIRETO DIRETOR",
            "Função": "ASSESSOR",
            "NOME_GESTOR": "",
            "Gerente": "",
            "Diretor/Sócio": "MARINO PEDRESCHI JUNIOR",
            "Status": "Ativo",
            "FERIAS": "Não",
        },
        {
            "Descrição": "OUTRO",
            "Empregado": "6",
            "Nome": "FORA DA ARVORE",
            "Função": "ANALISTA",
            "NOME_GESTOR": "OUTRO",
            "Gerente": "OUTRO",
            "Diretor/Sócio": "OUTRO DIRETOR",
            "Status": "Ativo",
            "FERIAS": "Não",
        },
    ]
    base = preparar_base_filtros(pd.DataFrame(linhas))
    filtrada = aplicar_filtros_consulta(
        base, {"diretor_socio": ["MARINO PEDRESCHI JUNIOR"]}
    )
    nomes = set(filtrada["Nome"].tolist())
    assert nomes == {
        "MARINO PEDRESCHI JUNIOR",
        "GERENTE SOB DIRETOR",
        "GESTOR SOB GERENTE",
        "COLAB SOB GESTOR",
        "COLAB DIRETO DIRETOR",
    }
    assert "FORA DA ARVORE" not in nomes
    assert filtrada["Empregado"].is_unique
    assert "Diretor/Sócio" in preparar_registros_setor(filtrada)[0]
